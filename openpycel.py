#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import os

from openpyxl import load_workbook
from pycel import ExcelCompiler


def openpycel():
    # path = os.path.dirname(__file__) #get current file path
    download_folder = "C:\\Users\\sadaco\\Downloads"
    spy_path = os.path.join(download_folder, "^DJI.csv")
    nk_path = os.path.join(download_folder, "t1570.csv")
    # dollar-yen-exchange-rate-historical-chart.csv | pound-japanese-yen-exchange-rate-historical-chart.csv
    fxy_path = os.path.join(
        download_folder, "pound-japanese-yen-exchange-rate-historical-chart.csv"
    )
    xlsx_path = "T:\\mydocs\\BCPad\\data\\nc225.xlsx"

    book = load_workbook(filename=xlsx_path, data_only=False)  # data_only=Trueで数式削除される
    sheet = book["UPRO"]  # シート変更
    sheet.delete_rows(idx=1, amount=1024)  # 範囲削除

    with open(spy_path) as f:
        reader = csv.reader(f)
        for row in reader:
            # print(row)
            sheet.append(row)

    sheet = book["nikkei"]  # シート変更
    sheet.delete_rows(idx=1, amount=1024)  # 範囲削除

    with open(nk_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            sheet.append(row)

    sheet = book["usd"]  # シート変更
    sheet.delete_rows(idx=1, amount=1024)  # 範囲削除

    with open(fxy_path) as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i < 7200: #12120
                continue
            else:
                sheet.append(row)

    MAX_RANGE = len(book["data"]["A"])

    book.save(xlsx_path)  # over write save
    book.close()  # Only affects read_only and write_only

    excel = ExcelCompiler(filename=xlsx_path)  # raspberry piでエクセルを開かずに再計算するための最終兵器
    book = load_workbook(filename=xlsx_path)
    sheet = book["data"]  # シート変更

    for i in range(MAX_RANGE):
        cell_C = f"C{i + 2}"  # upro
        cell_E = f"E{i + 2}"  # fxy
        cell_G = f"G{i + 2}"  # t1570
        # cell_J = f'J{i + 2}'

        sheet[f"B{i + 2}"] = excel.evaluate(cell_C)
        sheet[f"D{i + 2}"] = excel.evaluate(cell_E)
        sheet[f"F{i + 2}"] = excel.evaluate(cell_G)
        # sheet[f'I{i + 2}'] = excel.evaluate(cell_J)

    book.save(xlsx_path)
    book.close()
    print("Done openpycel")


if __name__ == "__main__":
    openpycel()
